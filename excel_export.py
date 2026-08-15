import io
import re
import urllib.parse
from copy import copy
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional


SOURCE_NAMES = {
    "github": "GitHub",
    "gitee": "Gitee",
    "gitlab": "GitLab",
    "huggingface": "Hugging Face",
    "stackoverflow": "Stack Overflow",
}

CANDIDATE_HEADERS = [
    "ID", "候选人", "账号", "来源", "城市", "建议岗位", "匹配评分", "单位/院校",
    "公开学历线索", "学历核验", "年龄核验", "工作地点核验", "Agent 项目核验",
    "审核状态", "联系进度", "联系进度更新时间", "联系方式等级", "公开邮箱",
    "邮箱来源", "邮箱核验时间", "公开主页", "其他公开入口", "公开简介", "审核备注",
    "首次采集", "最近采集", "记录更新时间",
]

EVIDENCE_HEADERS = [
    "候选人 ID", "候选人", "项目名称", "项目链接", "语言", "Stars", "项目类型", "公开说明",
]

CANDIDATE_WIDTHS = {
    "A": 8, "B": 18, "C": 17, "D": 10, "E": 10, "F": 20, "G": 11, "H": 22,
    "I": 25, "J": 15, "K": 15, "L": 18, "M": 20, "N": 14, "O": 14, "P": 20,
    "Q": 15, "R": 25, "S": 42, "T": 20, "U": 42, "V": 46, "W": 38, "X": 30,
    "Y": 20, "Z": 20, "AA": 20,
}

EVIDENCE_WIDTHS = {
    "A": 12, "B": 18, "C": 28, "D": 50, "E": 14, "F": 10, "G": 14, "H": 48,
}

SHANGHAI_TIMEZONE = timezone(timedelta(hours=8))
DATETIME_FORMAT = "yyyy-mm-dd hh:mm"
ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
UNSAFE_URL_CHARACTERS = re.compile(r"[\x00-\x20\x7f]")


def _safe_external_url(value: Any) -> str:
    url = str(value or "").strip()
    if not url or UNSAFE_URL_CHARACTERS.search(url):
        return ""
    try:
        parsed = urllib.parse.urlparse(url)
        hostname = parsed.hostname
        parsed.port
    except ValueError:
        return ""
    return url if parsed.scheme in {"http", "https"} and hostname else ""


def _safe_mailto(value: Any) -> str:
    email = str(value or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", email):
        return ""
    if "noreply" in email.casefold():
        return ""
    return "mailto:" + email


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, str):
        return "'" + value if value.startswith(("=", "+", "-", "@", "\t", "\r", "\n")) else value
    if isinstance(value, list):
        return [_excel_safe_value(item) for item in value]
    if isinstance(value, dict):
        return {key: _excel_safe_value(item) for key, item in value.items()}
    return value


def _excel_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = ILLEGAL_XML_CHARACTERS.sub("", text)
    return str(_excel_safe_value(text))


def _excel_integer(value: Any) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError, OverflowError):
        return 0


def _excel_shanghai_datetime(value: Any) -> Any:
    if value in (None, ""):
        return ""
    parsed: Optional[datetime]
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        raw = str(value).strip()
        if not raw:
            return ""
        try:
            parsed = datetime.fromisoformat(raw[:-1] + "+00:00" if raw.endswith(("Z", "z")) else raw)
        except ValueError:
            return _excel_text(value)
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(SHANGHAI_TIMEZONE).replace(tzinfo=None)
    return parsed


def _require_openpyxl() -> Dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import DataBarRule, FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Excel 导出需要 openpyxl 运行环境") from exc
    return {
        "Workbook": Workbook,
        "DataBarRule": DataBarRule,
        "FormulaRule": FormulaRule,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "DataValidation": DataValidation,
        "Table": Table,
        "TableStyleInfo": TableStyleInfo,
    }


def _candidate_row(candidate: Dict[str, Any]) -> List[Any]:
    source = _excel_text(candidate.get("source"))
    profile_url = _excel_text(candidate.get("profile_url"))
    contact_url = _excel_text(candidate.get("contact_url"))
    return [
        _excel_integer(candidate.get("id")),
        _excel_text(candidate.get("display_name")),
        _excel_text(candidate.get("username")),
        SOURCE_NAMES.get(source, source.upper()),
        _excel_text(candidate.get("city")),
        _excel_text(candidate.get("suggested_role")),
        _excel_integer(candidate.get("match_score")),
        _excel_text(candidate.get("company")) or "待核验",
        _excel_text(candidate.get("education_status")),
        _excel_text(candidate.get("education_verification")) or "待本人确认",
        _excel_text(candidate.get("age_status")) or "待本人确认",
        _excel_text(candidate.get("work_location_status")) or "待本人确认",
        _excel_text(candidate.get("agent_experience_status")) or "待人工核验",
        _excel_text(candidate.get("review_status")),
        _excel_text(candidate.get("contact_stage")) or "未联系",
        _excel_shanghai_datetime(candidate.get("contact_updated_at")),
        _excel_text(candidate.get("contact_level")) or "D",
        _excel_text(candidate.get("contact_email")),
        _excel_text(candidate.get("contact_email_source_url")),
        _excel_shanghai_datetime(candidate.get("contact_email_verified_at")),
        profile_url,
        contact_url if contact_url and contact_url != profile_url else "",
        _excel_text(candidate.get("bio")),
        _excel_text(candidate.get("review_note")),
        _excel_shanghai_datetime(candidate.get("first_seen_at")),
        _excel_shanghai_datetime(candidate.get("last_seen_at")),
        _excel_shanghai_datetime(candidate.get("updated_at")),
    ]


def _set_hyperlink(cell: Any, target: str, label: str) -> None:
    if not target:
        return
    cell.value = _excel_text(label)
    cell.hyperlink = target
    font = copy(cell.font)
    font.color = "0563C1"
    font.underline = "single"
    cell.font = font


def _apply_candidate_links(sheet: Any, candidates: List[Dict[str, Any]]) -> None:
    for row_index, candidate in enumerate(candidates, start=7):
        email_url = _safe_mailto(candidate.get("contact_email"))
        if email_url:
            _set_hyperlink(sheet.cell(row_index, 18), email_url, email_url[7:])
        email_source_url = _safe_external_url(candidate.get("contact_email_source_url"))
        if email_source_url:
            _set_hyperlink(sheet.cell(row_index, 19), email_source_url, email_source_url)
        profile_url = _safe_external_url(candidate.get("profile_url"))
        if profile_url:
            _set_hyperlink(sheet.cell(row_index, 21), profile_url, profile_url)
        contact_url = _safe_external_url(candidate.get("contact_url"))
        if contact_url and contact_url != profile_url:
            _set_hyperlink(sheet.cell(row_index, 22), contact_url, contact_url)


def _add_list_validation(sheet: Any, cell_range: str, values: List[str], data_validation: Any) -> None:
    validation = data_validation(
        type="list",
        formula1='"{}"'.format(",".join(values)),
        allow_blank=True,
    )
    validation.error = "请选择列表中的值"
    validation.errorTitle = "无效值"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _style_candidate_sheet(sheet: Any, candidates: List[Dict[str, Any]], modules: Dict[str, Any]) -> None:
    Alignment = modules["Alignment"]
    Border = modules["Border"]
    DataBarRule = modules["DataBarRule"]
    DataValidation = modules["DataValidation"]
    Font = modules["Font"]
    FormulaRule = modules["FormulaRule"]
    PatternFill = modules["PatternFill"]
    Side = modules["Side"]
    Table = modules["Table"]
    TableStyleInfo = modules["TableStyleInfo"]

    dark_fill = PatternFill("solid", fgColor="17343B")
    teal_fill = PatternFill("solid", fgColor="087F73")
    pale_teal_fill = PatternFill("solid", fgColor="E8F3F1")
    warning_fill = PatternFill("solid", fgColor="FFF4DF")
    white_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="E1E8EA")

    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "北京 / 重庆 AI 候选人总表"
    sheet.merge_cells("A2:N2")
    sheet["A2"] = "本地导出于 {} · 联系前请再次核验公开资料".format(
        datetime.now(SHANGHAI_TIMEZONE).strftime("%Y/%m/%d %H:%M:%S")
    )
    sheet["A3"], sheet["C3"], sheet["E3"], sheet["G3"] = (
        "候选人总数", "已审核", "已联系", "进入面试",
    )
    last_row = max(7, 6 + len(candidates))
    sheet["B3"] = "=COUNTA(B7:B{})".format(last_row)
    sheet["D3"] = '=COUNTIF(N7:N{},"<>待审核")'.format(last_row)
    sheet["F3"] = '=COUNTIF(O7:O{},"<>未联系")'.format(last_row)
    sheet["H3"] = '=COUNTIF(O7:O{},"进入面试")'.format(last_row)
    sheet.merge_cells("A5:N5")
    sheet["A5"] = "核验字段与联系进度可在人才雷达中维护；年龄、学历和到岗意愿不得仅凭公开资料推断。"
    sheet.append([])
    for column_index, header in enumerate(CANDIDATE_HEADERS, start=1):
        sheet.cell(6, column_index, header)
    for candidate in candidates:
        sheet.append(_candidate_row(candidate))

    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=27):
        for cell in row:
            cell.fill = dark_fill
            cell.font = Font(color="FFFFFF", bold=True, size=18)
            cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    for cell in sheet[2][:27]:
        cell.fill = pale_teal_fill
        cell.font = Font(color="52636A")
    for cell in sheet[3][:8]:
        cell.font = Font(color="087F73", bold=True)
        cell.border = Border(top=thin_side, bottom=thin_side)
    sheet["A3"].border = Border(left=thin_side, top=thin_side, bottom=thin_side)
    sheet["H3"].border = Border(right=thin_side, top=thin_side, bottom=thin_side)
    for cell in sheet[5][:27]:
        cell.fill = warning_fill
        cell.font = Font(color="86521D")
        cell.alignment = Alignment(wrap_text=True)
    for cell in sheet[6][:27]:
        cell.fill = teal_fill
        cell.font = white_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[6].height = 34

    if candidates:
        for row in sheet.iter_rows(min_row=7, max_row=last_row, min_col=1, max_col=27):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (23, 24))
                cell.border = Border(bottom=thin_side)
        for column in (16, 20, 25, 26, 27):
            for row_index in range(7, last_row + 1):
                sheet.cell(row_index, column).number_format = DATETIME_FORMAT
        for row_index in range(7, last_row + 1):
            sheet.cell(row_index, 7).number_format = "0"

        _add_list_validation(sheet, "J7:J{}".format(last_row), [
            "待本人确认", "本科及以上", "不符合",
        ], DataValidation)
        _add_list_validation(sheet, "K7:K{}".format(last_row), [
            "待本人确认", "30岁以下", "不符合",
        ], DataValidation)
        _add_list_validation(sheet, "L7:L{}".format(last_row), [
            "待本人确认", "接受北京", "接受重庆", "接受北京/重庆", "不接受",
        ], DataValidation)
        _add_list_validation(sheet, "M7:M{}".format(last_row), [
            "待人工核验", "原创 Agent 项目", "参与 Agent 项目", "仅关键词命中", "无相关经验",
        ], DataValidation)
        _add_list_validation(sheet, "N7:N{}".format(last_row), [
            "待审核", "优先联系", "需要核验", "人才储备", "不符合",
        ], DataValidation)
        _add_list_validation(sheet, "O7:O{}".format(last_row), [
            "未联系", "已联系", "已回复", "进入面试", "已发 Offer", "已录用", "不再推进",
        ], DataValidation)
        _add_list_validation(sheet, "Q7:Q{}".format(last_row), [
            "A", "B", "C", "D",
        ], DataValidation)

        sheet.conditional_formatting.add(
            "G7:G{}".format(last_row),
            DataBarRule(start_type="min", end_type="max", color="087F73"),
        )
        sheet.conditional_formatting.add(
            "N7:N{}".format(last_row),
            FormulaRule(
                formula=['NOT(ISERROR(SEARCH("不符合",N7)))'],
                fill=PatternFill("solid", fgColor="FFF0EE"),
                font=Font(color="B42318"),
            ),
        )
        sheet.conditional_formatting.add(
            "O7:O{}".format(last_row),
            FormulaRule(
                formula=['NOT(ISERROR(SEARCH("进入面试",O7)))'],
                fill=PatternFill("solid", fgColor="E8F6F3"),
                font=Font(color="08665D", bold=True),
            ),
        )
        sheet.conditional_formatting.add(
            "Q7:Q{}".format(last_row),
            FormulaRule(
                formula=['$Q7="A"'],
                fill=PatternFill("solid", fgColor="E8F6F3"),
                font=Font(color="08665D", bold=True),
            ),
        )
        table = Table(displayName="CandidatesTable", ref="A6:AA{}".format(last_row))
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    _apply_candidate_links(sheet, candidates)
    sheet.freeze_panes = "C7"
    for column, width in CANDIDATE_WIDTHS.items():
        sheet.column_dimensions[column].width = width


def _style_evidence_sheet(sheet: Any, candidates: List[Dict[str, Any]], modules: Dict[str, Any]) -> None:
    Alignment = modules["Alignment"]
    Border = modules["Border"]
    Font = modules["Font"]
    PatternFill = modules["PatternFill"]
    Side = modules["Side"]
    Table = modules["Table"]
    TableStyleInfo = modules["TableStyleInfo"]

    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "候选人公开项目证据"
    sheet.merge_cells("A2:H2")
    sheet["A2"] = "仅收录公开项目链接；原创/Fork 状态仍需结合仓库贡献记录核验。"
    for column_index, header in enumerate(EVIDENCE_HEADERS, start=1):
        sheet.cell(3, column_index, header)

    evidence_count = 0
    for candidate in candidates:
        for evidence in candidate.get("evidence") or []:
            evidence_count += 1
            sheet.append([
                _excel_integer(candidate.get("id")),
                _excel_text(candidate.get("display_name")),
                _excel_text(evidence.get("title")),
                _excel_text(evidence.get("url")),
                _excel_text(evidence.get("language")),
                _excel_integer(evidence.get("stars")),
                "Fork" if evidence.get("is_fork") else "原创仓库",
                _excel_text(evidence.get("description")),
            ])
            url = _safe_external_url(evidence.get("url"))
            if url:
                _set_hyperlink(sheet.cell(3 + evidence_count, 4), url, url)

    dark_fill = PatternFill("solid", fgColor="17343B")
    teal_fill = PatternFill("solid", fgColor="087F73")
    warning_fill = PatternFill("solid", fgColor="FFF4DF")
    thin_side = Side(style="thin", color="E1E8EA")
    for cell in sheet[1][:8]:
        cell.fill = dark_fill
        cell.font = Font(color="FFFFFF", bold=True, size=18)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    for cell in sheet[2][:8]:
        cell.fill = warning_fill
        cell.font = Font(color="86521D")
    for cell in sheet[3][:8]:
        cell.fill = teal_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)

    last_row = max(4, 3 + evidence_count)
    if evidence_count:
        for row in sheet.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=8):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=cell.column == 8)
                cell.border = Border(bottom=thin_side)
        for row_index in range(4, last_row + 1):
            sheet.cell(row_index, 6).number_format = "0"
        table = Table(displayName="EvidenceTable", ref="A3:H{}".format(last_row))
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    sheet.freeze_panes = "A4"
    for column, width in EVIDENCE_WIDTHS.items():
        sheet.column_dimensions[column].width = width


def generate_excel(
    candidates: List[Dict[str, Any]], preview_dir: Optional[Path] = None
) -> bytes:
    modules = _require_openpyxl()
    workbook = modules["Workbook"]()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    candidates_sheet = workbook.create_sheet("候选人总表")
    evidence_sheet = workbook.create_sheet("项目证据")

    _style_candidate_sheet(candidates_sheet, candidates, modules)
    _style_evidence_sheet(evidence_sheet, candidates, modules)
    workbook.active = 0
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"

    # Kept for API compatibility. The portable exporter does not require a renderer.
    _ = preview_dir
    output = io.BytesIO()
    try:
        workbook.save(output)
        return output.getvalue()
    except (OSError, ValueError, TypeError) as exc:
        raise RuntimeError("Excel 导出生成失败") from exc
    finally:
        workbook.close()
