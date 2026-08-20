import importlib.util
import io
import os
import re
import http.client
import json
import sqlite3
import tempfile
import threading
import unittest
import urllib.error
from datetime import datetime
from pathlib import Path
from unittest.mock import Mock, patch

import collectors
import contactability
import db
import excel_export
import app
import jobs
import source_health
from http.server import ThreadingHTTPServer
from jobs import JobManager, format_retry_time, next_weekly_run, normalize_config
from report import generate_report
from scoring import city_matches, detect_city, score_candidate


class FrontendStructureTests(unittest.TestCase):
    def test_javascript_id_references_exist_once_in_html(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        html = (project_dir / "static" / "index.html").read_text(encoding="utf-8")
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")
        html_ids = re.findall(r'\bid=["\']([^"\']+)["\']', html)
        javascript_ids = set(re.findall(r'\$\("#([A-Za-z][A-Za-z0-9_-]*)"\)', script))

        duplicates = sorted({item for item in html_ids if html_ids.count(item) > 1})
        missing = sorted(javascript_ids.difference(html_ids))

        self.assertEqual(duplicates, [])
        self.assertEqual(missing, [])

    def test_file_preview_loads_local_assets_and_shows_service_hint(self) -> None:
        static_dir = Path(__file__).resolve().parents[1] / "static"
        html = (static_dir / "index.html").read_text(encoding="utf-8")
        script = (static_dir / "app.js").read_text(encoding="utf-8")

        self.assertIn('href="styles.css"', html)
        self.assertIn('src="app.js"', html)
        self.assertTrue((static_dir / "styles.css").is_file())
        self.assertTrue((static_dir / "app.js").is_file())
        self.assertIn('window.location.protocol === "file:"', script)
        self.assertIn("http://127.0.0.1:8765/", script)

    def test_candidate_pool_displays_collection_times(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        html = (project_dir / "static" / "index.html").read_text(encoding="utf-8")
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")
        self.assertIn("<th>最近采集</th>", html)
        self.assertIn("candidate.last_seen_at", script)
        self.assertIn("candidate.first_seen_at", script)

    def test_candidate_pool_exposes_date_filter_controls_and_responsive_row(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        html = (project_dir / "static" / "index.html").read_text(encoding="utf-8")
        styles = (project_dir / "static" / "styles.css").read_text(encoding="utf-8")

        for element_id in (
            "candidateDateRange",
            "candidateDateFrom",
            "candidateDateTo",
            "candidateDateClearButton",
            "candidateFilterButton",
            "candidateExportButton",
        ):
            self.assertEqual(html.count('id="{}"'.format(element_id)), 1)

        date_filter = re.search(
            r'<div class="candidate-date-filter"[^>]*>(.*?)</div>', html, re.DOTALL
        )
        self.assertIsNotNone(date_filter)
        markup = date_filter.group(1)
        self.assertIn('id="candidateFilterButton"', markup)
        self.assertRegex(
            html,
            r'<a id="candidateExportButton"[^>]*href="/export/candidates\.xlsx"[^>]*>导出 Excel</a>',
        )
        for value in ("all", "today", "this_week", "last_7_days", "last_30_days", "this_month", "custom"):
            self.assertIn('value="{}"'.format(value), markup)
        for element_id in ("candidateDateFrom", "candidateDateTo"):
            self.assertRegex(
                markup,
                r'<input id="{}" type="date" min="2000-01-01" max="2100-12-31">'.format(element_id),
            )

        self.assertLess(html.index('src="candidate_date_filters.js"'), html.index('src="app.js"'))
        self.assertIn(".candidate-date-filter", styles)
        self.assertRegex(
            styles,
            r"\.candidate-date-filter\s*\{[^}]*grid-template-columns:\s*repeat\(3, minmax\(150px, 1fr\)\) repeat\(2, auto\)",
        )
        self.assertRegex(
            styles,
            r"\.candidate-date-filter label\s*\{[^}]*min-width:\s*0",
        )
        tablet_styles = re.search(r"@media \(max-width: 980px\) \{(.*?)\n\}", styles, re.DOTALL)
        self.assertIsNotNone(tablet_styles)
        self.assertIn(".candidate-date-filter", tablet_styles.group(1))
        self.assertIn("repeat(2, minmax(0, 1fr))", tablet_styles.group(1))
        mobile_styles = re.search(r"@media \(max-width: 640px\) \{(.*?)\n\}", styles, re.DOTALL)
        self.assertIsNotNone(mobile_styles)
        self.assertIn(".candidate-date-filter", mobile_styles.group(1))
        self.assertIn("grid-template-columns: 1fr", mobile_styles.group(1))

    def test_candidate_filters_keep_independent_draft_and_applied_state(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")

        for helper in (
            "candidateDateRangeForPreset",
            "candidateDateValidationMessage",
            "candidateExportUrl",
            "commitCandidatePageAfterFetch",
            "formatCandidateTime",
            "normalizeCandidateDateRange",
        ):
            self.assertIn(helper, script)
        for function_name in (
            "defaultCandidateFilters",
            "readCandidateFilterControls",
            "normalizeCandidateFilters",
            "writeCandidateDateControls",
            "buildCandidateListParams",
            "hasActiveCandidateFilters",
            "updateCandidateExportButton",
            "syncCandidateDraft",
        ):
            self.assertIn("function {}(".format(function_name), script)

        self.assertIn("candidateFilters: {", script)
        self.assertIn("draft: defaultCandidateFilters()", script)
        self.assertIn("applied: defaultCandidateFilters()", script)
        for field in (
            "search",
            "status",
            "city",
            "source",
            "contactability",
            "contact_stage",
            "date_range",
            "last_seen_from",
            "last_seen_to",
        ):
            self.assertIn("{}:".format(field), script)
        self.assertIn("candidateExportUrl(state.candidateFilters.applied)", script)

    def test_candidate_page_commits_filters_and_pagination_only_after_fetch(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")

        for function_name in (
            "fetchCandidatePage",
            "commitCandidatePage",
            "loadCandidatePage",
            "loadCandidates",
            "applyCandidateFilters",
        ):
            self.assertIn("function {}(".format(function_name), script)
        self.assertIn("commitCandidatePageAfterFetch(", script)
        self.assertIn("state.candidateFilters.applied = { ...filters }", script)
        self.assertIn("candidateDateValidationMessage(filters)", script)
        self.assertIn("candidateRequestId: 0", script)
        self.assertIn("const requestId = ++state.candidateRequestId", script)
        self.assertIn(
            "const isCurrent = () => requestId === state.candidateRequestId",
            script,
        )
        self.assertRegex(
            script,
            r"commitCandidatePageAfterFetch\([\s\S]*?isCurrent,\s*\);",
        )
        self.assertIn("last_seen_from", script)
        self.assertIn("last_seen_to", script)
        self.assertRegex(
            script,
            r"if \(normalized\.last_seen_from\) params\.set\(\"last_seen_from\"",
        )
        self.assertRegex(
            script,
            r"if \(normalized\.last_seen_to\) params\.set\(\"last_seen_to\"",
        )
        self.assertIn("const targetOffset = Math.max(0,", script)
        self.assertNotIn(
            "state.candidatePagination.offset += state.candidatePagination.limit",
            script,
        )

    def test_candidate_filter_events_empty_state_and_times_use_applied_filters(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")

        self.assertIn('cell.textContent = hasActiveCandidateFilters()', script)
        self.assertIn('"当前筛选条件下暂无候选人"', script)
        self.assertIn('"暂无候选人"', script)
        self.assertIn(
            "formatCandidateTime(candidate.last_seen_at)",
            script,
        )
        self.assertIn(
            "formatCandidateTime(candidate.first_seen_at)",
            script,
        )
        self.assertIn(
            'addDetail(details, "来源更新时间", formatCandidateTime(candidate.source_updated_at))',
            script,
        )
        self.assertIn(
            '$("#candidateFilterButton").addEventListener("click", applyCandidateFilters)',
            script,
        )
        self.assertIn(
            'if (event.key === "Enter") applyCandidateFilters();',
            script,
        )
        self.assertIn('$("#candidateDateRange").addEventListener("change"', script)
        self.assertIn("candidateDateRangeForPreset(event.currentTarget.value)", script)
        self.assertIn('$("#candidateDateRange").value = "custom"', script)
        self.assertIn('$("#candidateDateClearButton").addEventListener("click"', script)
        self.assertIn('writeCandidateDateControls(candidateDateRangeForPreset("all"))', script)
        self.assertIn("syncCandidateDraft();", script)

    def test_manual_modes_have_independent_default_on_ai_controls(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        html = (project_dir / "static" / "index.html").read_text(encoding="utf-8")
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")
        url_start = html.index('<div id="urlFields"')
        url_end = html.index("</section>", url_start)
        url_markup = html[url_start:url_end]
        self.assertIn('id="manualEnableAI" type="checkbox" checked', html)
        self.assertIn('id="manualUrlEnableAI" type="checkbox" checked', url_markup)
        self.assertIn(
            'state.manualMode === "url" ? $("#manualUrlEnableAI") : $("#manualEnableAI")',
            script,
        )
        self.assertIn("enable_ai: aiEnabled", script)
        self.assertIn("use_local_ai: aiEnabled", script)

    def test_candidate_ai_status_shows_fallback_reason_separately(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        html = (project_dir / "static" / "index.html").read_text(encoding="utf-8")
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")
        styles = (project_dir / "static" / "styles.css").read_text(encoding="utf-8")
        self.assertIn("function aiStatusText(candidate)", script)
        self.assertIn("function aiStatusReason(candidate)", script)
        self.assertIn('"不可用，已回退规则"', script)
        self.assertIn("ai_match_reason", script)
        self.assertIn("score-ai-reason", script)
        self.assertIn("状态：AI 未启用", script)
        self.assertIn("原因", script)
        self.assertIn(".score-ai-reason", styles)
        self.assertIn('id="dialogAiAnalysis"', html)

    def test_candidate_pool_exposes_batch_ai_reanalysis_controls(self) -> None:
        project_dir = Path(__file__).resolve().parents[1]
        html = (project_dir / "static" / "index.html").read_text(encoding="utf-8")
        script = (project_dir / "static" / "app.js").read_text(encoding="utf-8")
        styles = (project_dir / "static" / "styles.css").read_text(encoding="utf-8")
        self.assertIn('id="reanalyzeAiButton"', html)
        self.assertIn('id="candidateAiProgress"', html)
        self.assertIn('id="candidateAiProgressFill"', html)
        self.assertIn('/api/candidates/reanalyze-ai', script)
        self.assertIn('/api/jobs/${encodeURIComponent(jobId)}', script)
        self.assertIn("error.status = response.status", script)
        self.assertIn("if (error.status !== 404) throw error", script)
        self.assertIn('ai_match_status', script)
        self.assertIn('.ai-batch-progress', styles)


class SourceHealthTests(unittest.TestCase):
    def test_static_source_health_distinguishes_supported_and_planned_sources(self) -> None:
        payload = source_health.get_source_health()
        items = {item["key"]: item for item in payload["items"]}
        self.assertEqual(items["github"]["status"], "not_checked")
        self.assertEqual(items["modelscope"]["status"], "not_implemented")
        self.assertEqual(items["douyin"]["status"], "manual_only")
        self.assertEqual(len(items), 15)

    def test_probe_maps_network_failure_without_reading_candidate_data(self) -> None:
        source = next(item for item in source_health.SOURCE_CATALOG if item["key"] == "github")
        with patch.object(
            source_health.urllib.request,
            "urlopen",
            side_effect=source_health.urllib.error.URLError("offline"),
        ):
            result = source_health._probed_result(source)
        self.assertEqual(result["status"], "network_unavailable")
        self.assertIn("无法连接", result["detail"])

class ScoringTests(unittest.TestCase):
    def test_city_detection_handles_chinese_and_english(self) -> None:
        self.assertEqual(detect_city("Beijing, China"), "北京")
        self.assertEqual(detect_city("重庆大学"), "重庆")
        self.assertEqual(detect_city("Shanghai"), "待核验")

    def test_original_agent_projects_raise_score(self) -> None:
        profile = {
            "display_name": "Candidate",
            "username": "candidate",
            "bio": "AI agent engineer in Beijing",
            "company": "Example",
            "city": "北京",
            "contact_email": "candidate@example.com",
            "source_updated_at": "2026-07-01",
        }
        evidence = [
            {
                "title": "agent-reliability-kit",
                "description": "MCP tool calling and agent evaluation",
                "stars": 30,
                "is_fork": False,
            }
        ]
        score, role, ranked = score_candidate(profile, evidence, "AI Agent 工程师", "北京")
        self.assertGreaterEqual(score, 70)
        self.assertEqual(role, "AI Agent 工程师")
        self.assertEqual(ranked[0]["title"], "agent-reliability-kit")

    def test_public_email_does_not_change_match_score(self) -> None:
        profile = {
            "display_name": "Candidate",
            "username": "candidate",
            "bio": "AI agent engineer in Beijing",
            "company": "Example",
            "city": "北京",
            "source_updated_at": "2026-07-01",
        }
        evidence = [
            {
                "title": "agent-kit",
                "description": "MCP tool calling agent",
                "stars": 8,
                "is_fork": False,
            }
        ]
        without_email = score_candidate(
            profile, evidence, "AI Agent 工程师", "北京"
        )[0]
        with_email = score_candidate(
            {**profile, "contact_email": "candidate@example.test"},
            evidence,
            "AI Agent 工程师",
            "北京",
        )[0]
        self.assertEqual(with_email, without_email)


class ContactabilityTests(unittest.TestCase):
    def test_contact_levels_require_valid_public_evidence(self) -> None:
        profile_url = "https://profiles.example.test/candidate"
        verified = {
            "profile_url": profile_url,
            "contact_url": profile_url,
            "contact_email": "candidate@example.test",
            "contact_email_source_url": profile_url,
            "contact_email_verified_at": "2026-08-04T10:00:00+08:00",
        }
        self.assertEqual(contactability.derive_contact_level(verified), "A")
        self.assertEqual(
            contactability.derive_contact_level(
                {
                    **verified,
                    "contact_email_source_url": "",
                    "contact_email_verified_at": "",
                }
            ),
            "B",
        )
        self.assertEqual(
            contactability.derive_contact_level(
                {
                    "profile_url": profile_url,
                    "contact_url": "https://contact.example.test/form",
                }
            ),
            "C",
        )
        self.assertEqual(
            contactability.derive_contact_level(
                {
                    "profile_url": profile_url,
                    "contact_url": profile_url + "/?from=radar",
                    "contact_email": "123@users.noreply.github.com",
                }
            ),
            "D",
        )
        self.assertEqual(
            contactability.derive_contact_level(
                {
                    "profile_url": profile_url,
                    "contact_url": "javascript:alert(1)",
                }
            ),
            "D",
        )


class ScheduleTests(unittest.TestCase):
    def test_retry_time_format_does_not_depend_on_system_locale(self) -> None:
        self.assertEqual(
            format_retry_time("2026-08-15T09:05:00+08:00"),
            "08月15日 09:05",
        )

    def test_next_weekly_run_moves_to_next_week_after_time_passed(self) -> None:
        monday_after_run = datetime(2026, 7, 27, 10, 0, 0)
        result = next_weekly_run(0, 9, 0, monday_after_run)
        self.assertEqual(result, datetime(2026, 8, 3, 9, 0, 0))

    def test_normalize_config_caps_target_and_filters_values(self) -> None:
        config = normalize_config(
            {
                "target": 999,
                "roles": ["AI Coding 工程师", "无效岗位"],
                "cities": ["北京", "上海"],
                "sources": ["github", "unknown"],
            }
        )
        self.assertEqual(config["target"], 50)
        self.assertEqual(config["roles"], ["AI Coding 工程师"])
        self.assertEqual(config["cities"], ["北京"])
        self.assertEqual(config["sources"], ["github"])

    def test_normalize_config_accepts_new_public_sources(self) -> None:
        config = normalize_config(
            {
                "sources": ["gitlab", "huggingface", "stackoverflow"],
                "roles": ["AI Agent 工程师"],
                "cities": ["重庆"],
            }
        )
        self.assertEqual(
            config["sources"], ["gitlab", "huggingface", "stackoverflow"]
        )

    def test_contact_priority_defaults_on_and_requires_a_boolean(self) -> None:
        self.assertTrue(normalize_config({})["prefer_contactable"])
        self.assertFalse(
            normalize_config({"prefer_contactable": False})["prefer_contactable"]
        )
        with self.assertRaises(ValueError):
            normalize_config({"prefer_contactable": "false"})

    def test_normalize_config_rejects_unknown_mode(self) -> None:
        with self.assertRaises(ValueError):
            normalize_config({"mode": "unknown"})


class CollectorSourceTests(unittest.TestCase):
    def test_public_profile_email_prefers_explicit_mailto(self) -> None:
        markup = """
        <html><body>
          <a href="mailto:Candidate%40example.com?subject=Hello">Email</a>
          <p>support@example.org</p>
        </body></html>
        """
        self.assertEqual(
            collectors.extract_public_profile_email(markup),
            "candidate@example.com",
        )

    def test_public_profile_email_accepts_one_visible_address(self) -> None:
        self.assertEqual(
            collectors.extract_public_profile_email("<p>candidate@example.com</p>"),
            "candidate@example.com",
        )

    def test_public_profile_email_rejects_noreply_and_ambiguous_text(self) -> None:
        self.assertEqual(
            collectors.extract_public_profile_email(
                "<p>123@users.noreply.github.com</p>"
            ),
            "",
        )
        self.assertEqual(
            collectors.extract_public_profile_email(
                "<p>first@example.com and second@example.com</p>"
            ),
            "",
        )

    def test_public_profile_email_ignores_site_footer_contact(self) -> None:
        self.assertEqual(
            collectors.extract_public_profile_email(
                '<main><p>Public profile</p></main><footer><a href="mailto:support@example.com">Support</a></footer>'
            ),
            "",
        )

    def test_public_profile_email_ignores_nonsemantic_footer_contact(self) -> None:
        self.assertEqual(
            collectors.extract_public_profile_email(
                '<main><p>Public profile</p></main><div class="site-footer"><br/><a href="mailto:support@example.com">Support</a></div>'
            ),
            "",
        )

    def test_public_profile_email_rejects_multiple_mailto_addresses(self) -> None:
        self.assertEqual(
            collectors.extract_public_profile_email(
                '<a href="mailto:first@example.com">First</a><a href="mailto:second@example.com">Second</a>'
            ),
            "",
        )

    def test_shared_email_is_suppressed_only_within_one_source(self) -> None:
        candidates = [
            {"source": "github", "contact_email": "shared@example.com"},
            {"source": "github", "contact_email": "shared@example.com"},
            {"source": "gitee", "contact_email": "shared@example.com"},
        ]
        self.assertEqual(collectors.suppress_shared_public_emails(candidates), 2)
        self.assertEqual(candidates[0]["contact_email"], "")
        self.assertEqual(candidates[1]["contact_email"], "")
        self.assertEqual(candidates[2]["contact_email"], "shared@example.com")

    def test_fetch_json_retries_a_transient_network_error(self) -> None:
        class Response:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return b'{"ok": true}'

        with patch(
            "collectors.urllib.request.urlopen",
            side_effect=[urllib.error.URLError("offline"), Response()],
        ) as urlopen, patch("collectors.time.sleep"):
            self.assertEqual(collectors.fetch_json("https://example.com"), {"ok": True})
        self.assertEqual(urlopen.call_count, 2)

    def test_fetch_json_retries_an_incomplete_response(self) -> None:
        class Response:
            def __init__(self, broken=False):
                self.broken = broken

            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                if self.broken:
                    raise http.client.IncompleteRead(b"partial")
                return b'{"ok": true}'

        with patch(
            "collectors.urllib.request.urlopen",
            side_effect=[Response(broken=True), Response()],
        ) as urlopen, patch("collectors.time.sleep"):
            self.assertEqual(collectors.fetch_json("https://example.com"), {"ok": True})
        self.assertEqual(urlopen.call_count, 2)

    def test_huggingface_models_become_candidate_evidence(self) -> None:
        candidate = collectors.huggingface_candidate(
            "candidate",
            [
                {
                    "id": "candidate/agent-model",
                    "pipeline_tag": "text-generation",
                    "tags": ["agents", "tool-calling"],
                    "likes": 12,
                    "lastModified": "2026-07-20T10:00:00.000Z",
                }
            ],
            "AI Agent 工程师",
            "北京",
        )
        self.assertEqual(candidate["source"], "huggingface")
        self.assertEqual(candidate["city"], "待核验")
        self.assertEqual(candidate["age_status"], "待本人确认")
        self.assertEqual(candidate["evidence"][0]["title"], "agent-model")

    def test_gitlab_public_profile_keeps_explicit_public_contact(self) -> None:
        candidate = collectors.gitlab_candidate(
            {
                "id": 7,
                "username": "candidate",
                "name": "Candidate",
                "location": "Chongqing",
                "bio": "AI agent engineer",
                "organization": "Example University",
                "web_url": "https://gitlab.com/candidate",
                "website_url": "https://candidate.example.com",
                "public_email": "candidate@example.com",
            },
            [
                {
                    "name": "agent-kit",
                    "web_url": "https://gitlab.com/candidate/agent-kit",
                    "description": "MCP tool calling agent",
                    "star_count": 6,
                }
            ],
            "AI Agent 工程师",
            "重庆",
        )
        self.assertEqual(candidate["city"], "重庆")
        self.assertEqual(candidate["contact_email"], "candidate@example.com")
        self.assertEqual(candidate["contact_url"], "https://candidate.example.com")
        self.assertEqual(
            candidate["contact_email_source_url"],
            "https://gitlab.com/candidate",
        )
        self.assertTrue(candidate["contact_email_verified_at"])
        self.assertEqual(contactability.derive_contact_level(candidate), "A")

    def test_stackoverflow_profile_html_is_reduced_to_public_text(self) -> None:
        candidate = collectors.stackoverflow_candidate(
            {
                "user_id": 9,
                "display_name": "Candidate &amp; Builder",
                "location": "Beijing",
                "about_me": "<p>Building <strong>LLM agents</strong></p>",
                "link": "https://stackoverflow.com/users/9/candidate",
                "last_access_date": 1785000000,
            },
            [
                {
                    "title": "How to evaluate an LLM agent?",
                    "link": "https://stackoverflow.com/questions/1/example",
                    "tags": ["openai-api", "langchain"],
                    "score": 4,
                }
            ],
            "AI Agent 工程师",
            "北京",
        )
        self.assertEqual(candidate["display_name"], "Candidate & Builder")
        self.assertEqual(candidate["bio"], "Building LLM agents")
        self.assertEqual(candidate["city"], "北京")

    @patch("collectors.huggingface_user")
    def test_public_url_supports_huggingface(self, mocked_user) -> None:
        mocked_user.return_value = {"source": "huggingface"}
        result = collectors.analyze_public_url(
            "https://huggingface.co/candidate/agent-model",
            "AI Agent 工程师",
            "北京",
        )
        self.assertEqual(result["source"], "huggingface")
        mocked_user.assert_called_once_with("candidate", "AI Agent 工程师", "北京")

    @patch("collectors.github_user")
    @patch("collectors.fetch_json")
    def test_github_search_skips_one_invalid_public_account(
        self, mocked_fetch, mocked_user
    ) -> None:
        mocked_fetch.return_value = {"items": [{"login": "broken"}, {"login": "valid"}]}
        mocked_user.side_effect = [
            collectors.CollectorError("公开账号不存在"),
            {
                "source": "github",
                "external_id": "2",
                "username": "valid",
                "display_name": "Valid",
                "city": "北京",
                "profile_url": "https://github.com/valid",
                "contact_url": "https://github.com/valid",
                "evidence": [],
                "match_score": 50,
                "suggested_role": "AI Agent 工程师",
            },
        ]
        result = collectors.search_github("agent", "北京", "AI Agent 工程师", 2)
        self.assertEqual([item["username"] for item in result], ["valid"])


class DatabaseTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.previous_db = os.environ.get("TALENT_RADAR_DB")
        os.environ["TALENT_RADAR_DB"] = str(Path(self.temp_dir.name) / "test.db")
        db.init_db()

    def tearDown(self) -> None:
        if self.previous_db is None:
            os.environ.pop("TALENT_RADAR_DB", None)
        else:
            os.environ["TALENT_RADAR_DB"] = self.previous_db
        self.temp_dir.cleanup()

    def candidate(self, name: str = "First Name"):
        return {
            "source": "github",
            "external_id": "123",
            "username": "candidate",
            "display_name": name,
            "city": "北京",
            "company": "示例大学",
            "profile_url": "https://github.com/candidate",
            "contact_url": "https://candidate.example.com",
            "contact_email": "candidate@example.com",
            "suggested_role": "AI Agent 工程师",
            "match_score": 88,
            "evidence": [
                {
                    "title": "agent-kit",
                    "url": "https://github.com/candidate/agent-kit",
                    "description": "Agent tools",
                    "stars": 5,
                    "is_fork": False,
                }
            ],
        }

    def add_candidate_at(self, external_id, collected_at, **overrides):
        candidate = self.candidate(external_id)
        candidate.update(
            {
                "external_id": external_id,
                "username": external_id,
                "display_name": external_id,
                "profile_url": "https://profiles.example.test/{}".format(external_id),
                "contact_url": "https://profiles.example.test/{}".format(external_id),
                "contact_email": "{}@example.test".format(external_id),
                "evidence": [
                    {
                        "title": "{}-project".format(external_id),
                        "url": "https://code.example.test/{}/project".format(external_id),
                        "description": "synthetic agent project",
                        "stars": 1,
                        "is_fork": False,
                    }
                ],
            }
        )
        candidate.update(overrides)
        with patch.object(db, "now_iso", return_value=collected_at):
            candidate_id, _ = db.upsert_candidate(candidate)
        return candidate_id

    def enable_schedule(self) -> None:
        schedule = db.get_schedule()
        schedule["enabled"] = True
        db.save_schedule(schedule, "2099-01-01T09:00:00+08:00")

    def test_connection_context_closes_file_handle(self) -> None:
        connection = db.connect()
        with connection:
            connection.execute("SELECT 1")
        with self.assertRaises(sqlite3.ProgrammingError):
            connection.execute("SELECT 1")

    def test_schedule_retry_is_durable_and_capped(self) -> None:
        self.enable_schedule()
        first_retry = db.defer_schedule_retry(delay_minutes=0, max_retries=3)
        self.assertIsNotNone(first_retry)
        config = db.claim_due_schedule("2099-01-08T09:00:00+08:00")
        self.assertEqual(config["target"], 30)
        self.assertTrue(config["prefer_contactable"])
        self.assertIsNone(db.get_schedule()["retry_at"])

        self.assertIsNotNone(db.defer_schedule_retry(delay_minutes=0, max_retries=3))
        self.assertIsNotNone(db.defer_schedule_retry(delay_minutes=0, max_retries=3))
        self.assertIsNone(db.defer_schedule_retry(delay_minutes=0, max_retries=3))
        self.assertEqual(db.get_schedule()["retry_count"], 3)

        db.clear_schedule_retry()
        self.assertEqual(db.get_schedule()["retry_count"], 0)

    def test_automatic_network_failure_schedules_retry(self) -> None:
        self.enable_schedule()
        config = normalize_config(
            {
                "sources": ["github"],
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
            }
        )
        job_id = db.create_job("每周自动", config)
        manager = JobManager()
        with patch.object(
            manager,
            "_run_search_job",
            side_effect=collectors.NetworkUnavailable("offline"),
        ):
            manager._run_job(job_id, "每周自动", config)

        job = db.get_job(job_id)
        schedule = db.get_schedule()
        self.assertEqual(job["status"], "网络不可用")
        self.assertIn("自动补跑", job["message"])
        self.assertEqual(schedule["retry_count"], 1)
        self.assertIsNotNone(schedule["retry_at"])

    def test_upsert_deduplicates_by_source_and_external_id(self) -> None:
        first_id, inserted = db.upsert_candidate(self.candidate())
        second_id, second_inserted = db.upsert_candidate(self.candidate("Updated Name"))
        result = db.list_candidates({})
        self.assertTrue(inserted)
        self.assertFalse(second_inserted)
        self.assertEqual(first_id, second_id)
        self.assertEqual(result["total"], 1)
        self.assertEqual(result["items"][0]["display_name"], "Updated Name")

    def test_candidate_date_filter_validates_exact_supported_dates(self) -> None:
        cases = (
            ("last_seen_from", "2026-8-18", "开始日期格式无效，请使用 YYYY-MM-DD"),
            ("last_seen_from", "2026-02-30", "开始日期格式无效，请使用 YYYY-MM-DD"),
            ("last_seen_from", " 2026-08-18", "开始日期格式无效，请使用 YYYY-MM-DD"),
            ("last_seen_from", "２０２６-０８-１８", "开始日期格式无效，请使用 YYYY-MM-DD"),
            ("last_seen_from", "1999-12-31", "日期必须在 2000-01-01 至 2100-12-31 之间"),
            ("last_seen_to", "2101-01-01", "日期必须在 2000-01-01 至 2100-12-31 之间"),
        )
        for key, value, message in cases:
            with self.subTest(key=key, value=value):
                with self.assertRaisesRegex(ValueError, re.escape(message)):
                    db.list_candidates({key: value})

    def test_candidate_date_filter_rejects_reversed_range(self) -> None:
        with self.assertRaisesRegex(ValueError, re.escape("开始日期不能晚于结束日期")):
            db.list_candidates(
                {"last_seen_from": "2026-08-19", "last_seen_to": "2026-08-18"}
            )

    def test_candidate_date_filter_builds_beijing_time_boundaries(self) -> None:
        self.assertEqual(db._candidate_date_bounds({}), (None, None))
        self.assertEqual(
            db._candidate_date_bounds({"last_seen_from": "2026-08-18"}),
            ("2026-08-18T00:00:00+08:00", None),
        )
        self.assertEqual(
            db._candidate_date_bounds(
                {"last_seen_from": "2026-08-18", "last_seen_to": "2026-08-18"}
            ),
            ("2026-08-18T00:00:00+08:00", "2026-08-19T00:00:00+08:00"),
        )
        self.assertEqual(
            db._candidate_date_bounds({"last_seen_to": "2100-12-31"}),
            (None, "2101-01-01T00:00:00+08:00"),
        )

    def test_candidate_date_filter_uses_inclusive_beijing_boundaries(self) -> None:
        self.add_candidate_at("before", "2026-08-17T15:59:59+00:00")
        self.add_candidate_at("start", "2026-08-17T16:00:00+00:00")
        self.add_candidate_at("late", "2026-08-18T23:59:59.999999+08:00")
        self.add_candidate_at("next-day", "2026-08-19T00:00:00+08:00")

        result = db.list_candidates(
            {"last_seen_from": "2026-08-18", "last_seen_to": "2026-08-18"}
        )

        self.assertEqual({item["external_id"] for item in result["items"]}, {"start", "late"})
        self.assertEqual(result["total"], 2)

    def test_candidate_date_filter_supports_open_ended_ranges(self) -> None:
        self.add_candidate_at("old", "2026-08-10T10:00:00+08:00")
        self.add_candidate_at("new", "2026-08-20T10:00:00+08:00")

        from_result = db.list_candidates({"last_seen_from": "2026-08-18"})
        to_result = db.list_candidates({"last_seen_to": "2026-08-18"})

        self.assertEqual({item["external_id"] for item in from_result["items"]}, {"new"})
        self.assertEqual({item["external_id"] for item in to_result["items"]}, {"old"})

    def test_export_candidates_applies_date_filter_and_matching_evidence(self) -> None:
        self.add_candidate_at("in-range", "2026-08-18T09:00:00+08:00")
        self.add_candidate_at("out-of-range", "2026-08-17T09:00:00+08:00")

        result = db.export_candidates(
            {"last_seen_from": "2026-08-18", "last_seen_to": "2026-08-18"}
        )

        self.assertEqual([item["external_id"] for item in result], ["in-range"])
        self.assertEqual([item["title"] for item in result[0]["evidence"]], ["in-range-project"])

    def test_export_candidates_without_filters_remains_backward_compatible(self) -> None:
        self.add_candidate_at("first", "2026-08-17T09:00:00+08:00", match_score=70)
        self.add_candidate_at("second", "2026-08-18T09:00:00+08:00", match_score=90)

        result = db.export_candidates()

        self.assertEqual([item["external_id"] for item in result], ["second", "first"])
        self.assertEqual(
            [[evidence["title"] for evidence in item["evidence"]] for item in result],
            [["second-project"], ["first-project"]],
        )

    def test_export_candidates_rejects_invalid_and_reversed_date_filters(self) -> None:
        with self.assertRaisesRegex(ValueError, re.escape("开始日期格式无效，请使用 YYYY-MM-DD")):
            db.export_candidates({"last_seen_from": "2026-02-30"})
        with self.assertRaisesRegex(ValueError, re.escape("开始日期不能晚于结束日期")):
            db.export_candidates(
                {"last_seen_from": "2026-08-19", "last_seen_to": "2026-08-18"}
            )

    def test_export_candidates_reads_candidates_and_evidence_from_one_snapshot(self) -> None:
        candidate_id = self.add_candidate_at("in-range", "2026-08-18T09:00:00+08:00")
        archived = {"value": False}

        class ArchiveAfterCandidateSelectConnection(db.ClosingConnection):
            def execute(self, sql, parameters=()):
                cursor = super().execute(sql, parameters)
                if not archived["value"] and "SELECT * FROM candidates" in sql:
                    archived["value"] = True
                    with sqlite3.connect(str(db.db_path())) as writer:
                        writer.execute(
                            "UPDATE candidates SET archived_at = ? WHERE id = ?",
                            ("2026-08-18T10:00:00+08:00", candidate_id),
                        )
                return cursor

        with patch.object(db, "ClosingConnection", ArchiveAfterCandidateSelectConnection):
            result = db.export_candidates(
                {"last_seen_from": "2026-08-18", "last_seen_to": "2026-08-18"}
            )

        self.assertTrue(archived["value"])
        self.assertEqual([item["external_id"] for item in result], ["in-range"])
        self.assertEqual([item["title"] for item in result[0]["evidence"]], ["in-range-project"])

    def test_candidate_date_filter_combines_with_filters_and_pagination(self) -> None:
        self.add_candidate_at("beijing-a", "2026-08-18T09:00:00+08:00")
        self.add_candidate_at("beijing-b", "2026-08-18T10:00:00+08:00")
        self.add_candidate_at(
            "chongqing", "2026-08-18T11:00:00+08:00", city="重庆"
        )
        self.add_candidate_at(
            "wrong-source", "2026-08-18T11:30:00+08:00", source="gitee"
        )
        self.add_candidate_at(
            "profile-only",
            "2026-08-18T11:45:00+08:00",
            contact_email="",
            contact_url="https://profiles.example.test/profile-only",
        )
        self.add_candidate_at("outside", "2026-08-17T11:00:00+08:00")
        wrong_status = self.add_candidate_at("wrong-status", "2026-08-18T12:00:00+08:00")
        archived = self.add_candidate_at("archived", "2026-08-18T13:00:00+08:00")
        with db.connect() as connection:
            connection.execute(
                "UPDATE candidates SET review_status = ? WHERE id = ?",
                ("人才储备", wrong_status),
            )
            connection.execute(
                "UPDATE candidates SET archived_at = ? WHERE id = ?",
                ("2026-08-18T14:00:00+08:00", archived),
            )

        filters = {
            "last_seen_from": "2026-08-18",
            "last_seen_to": "2026-08-18",
            "city": "北京",
            "source": "github",
            "status": "待审核",
            "contactability": "email",
        }
        unpaginated = db.list_candidates(filters)
        self.assertEqual(unpaginated["total"], 2)
        self.assertEqual(
            {item["external_id"] for item in unpaginated["items"]},
            {"beijing-a", "beijing-b"},
        )

        result = db.list_candidates({**filters, "limit": 1, "offset": 1})

        self.assertEqual(result["total"], 2)
        self.assertEqual(len(result["items"]), 1)
        returned_ids = {item["external_id"] for item in result["items"]}
        self.assertTrue(returned_ids <= {"beijing-a", "beijing-b"})

    def test_review_and_report_keep_clickable_contacts(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        self.assertTrue(
            db.review_candidate(
                candidate_id,
                "优先联系",
                "项目证据充分",
                "本科及以上",
                "30岁以下",
                "接受北京",
                "原创 Agent 项目",
                "已联系",
            )
        )
        saved = db.get_candidate(candidate_id)
        self.assertEqual(saved["education_verification"], "本科及以上")
        self.assertEqual(saved["age_status"], "30岁以下")
        self.assertEqual(saved["work_location_status"], "接受北京")
        self.assertEqual(saved["agent_experience_status"], "原创 Agent 项目")
        self.assertEqual(saved["contact_stage"], "已联系")
        self.assertTrue(saved["contact_updated_at"])
        content = generate_report(db.report_candidates()).decode("utf-8")
        self.assertIn("mailto:candidate@example.com", content)
        self.assertIn("https://github.com/candidate", content)
        self.assertIn("公开主页", content)
        self.assertIn("https://candidate.example.com", content)
        self.assertIn("其他公开入口", content)
        self.assertIn("示例大学", content)
        self.assertIn("本科及以上", content)
        self.assertIn("原创 Agent 项目", content)
        self.assertIn("已联系", content)
        self.assertIn("优先联系", content)
        self.assertIn("匹配等级", content)
        self.assertIn("高匹配", content)
        self.assertIn("最近采集", content)
        self.assertIn("B级 · 公开邮箱待复核", content)
        self.assertNotIn("mail.google.com", content)

    def test_upsert_records_first_and_latest_collection_times(self) -> None:
        with patch.object(
            db,
            "now_iso",
            side_effect=[
                "2026-08-01T10:00:00+08:00",
                "2026-08-05T14:00:00+08:00",
            ],
        ):
            candidate_id, _ = db.upsert_candidate(self.candidate())
            first = db.get_candidate(candidate_id)
            db.upsert_candidate(self.candidate("Refreshed Name"))
            refreshed = db.get_candidate(candidate_id)
        self.assertTrue(first["first_seen_at"])
        self.assertEqual(first["first_seen_at"], "2026-08-01T10:00:00+08:00")
        self.assertEqual(first["last_seen_at"], first["first_seen_at"])
        self.assertEqual(refreshed["first_seen_at"], first["first_seen_at"])
        self.assertEqual(refreshed["last_seen_at"], "2026-08-05T14:00:00+08:00")

    def test_profile_refresh_without_ai_preserves_completed_ai_analysis(self) -> None:
        analyzed = self.candidate()
        analyzed.update(
            {
                "match_score": 90,
                "rule_match_score": 88,
                "ai_match_score": 92,
                "ai_match_status": "已完成",
                "ai_match_confidence": 0.91,
                "ai_match_reason": "公开项目与岗位模板高度相关",
                "ai_match_evidence": ["agent-kit"],
                "ai_match_model": "qwen3:4b",
                "ai_match_at": "2026-08-05T14:00:00+08:00",
            }
        )
        candidate_id, _ = db.upsert_candidate(analyzed)

        refreshed = self.candidate("Refreshed Name")
        db.upsert_candidate(refreshed)

        saved = db.get_candidate(candidate_id)
        self.assertEqual(saved["display_name"], "Refreshed Name")
        self.assertEqual(saved["match_score"], 90)
        self.assertEqual(saved["rule_match_score"], 88)
        self.assertEqual(saved["ai_match_score"], 92)
        self.assertEqual(saved["ai_match_status"], "已完成")
        self.assertEqual(saved["ai_match_confidence"], 0.91)
        self.assertEqual(saved["ai_match_reason"], "公开项目与岗位模板高度相关")
        self.assertEqual(saved["ai_match_evidence"], ["agent-kit"])
        self.assertEqual(saved["ai_match_model"], "qwen3:4b")
        self.assertEqual(saved["ai_match_at"], "2026-08-05T14:00:00+08:00")

    def test_manual_verification_survives_profile_refresh(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        db.review_candidate(
            candidate_id,
            "需要核验",
            "待沟通",
            "本科及以上",
            "30岁以下",
            "接受北京/重庆",
            "参与 Agent 项目",
            "已回复",
        )
        db.upsert_candidate(self.candidate("Refreshed Name"))
        saved = db.get_candidate(candidate_id)
        self.assertEqual(saved["display_name"], "Refreshed Name")
        self.assertEqual(saved["age_status"], "30岁以下")
        self.assertEqual(saved["education_verification"], "本科及以上")
        self.assertEqual(saved["contact_stage"], "已回复")

    def test_verified_public_email_survives_profile_refresh(self) -> None:
        candidate = self.candidate()
        candidate["contact_email"] = ""
        candidate_id, _ = db.upsert_candidate(candidate)
        self.assertTrue(
            db.set_public_email(
                candidate_id,
                "candidate@example.com",
                "https://github.com/candidate",
            )
        )
        candidate["display_name"] = "Refreshed Name"
        candidate["contact_email"] = "unverified@example.test"
        db.upsert_candidate(candidate)
        saved = db.get_candidate(candidate_id)
        self.assertEqual(saved["contact_email"], "candidate@example.com")
        self.assertEqual(
            saved["contact_email_source_url"],
            "https://github.com/candidate",
        )
        self.assertTrue(saved["contact_email_verified_at"])
        self.assertEqual(saved["contact_level"], "A")

    def test_init_db_migrates_and_backfills_contact_level(self) -> None:
        candidate = self.candidate()
        candidate["contact_email_source_url"] = candidate["profile_url"]
        candidate["contact_email_verified_at"] = "2026-08-04T10:00:00+08:00"
        candidate_id, _ = db.upsert_candidate(candidate)
        with db.connect() as connection:
            connection.execute("DROP INDEX idx_candidates_contact_level")
            connection.execute("ALTER TABLE candidates DROP COLUMN contact_level")

        db.init_db()

        with db.connect() as connection:
            columns = {
                row["name"]
                for row in connection.execute(
                    "PRAGMA table_info(candidates)"
                ).fetchall()
            }
        self.assertIn("contact_level", columns)
        self.assertEqual(db.get_candidate(candidate_id)["contact_level"], "A")

    def test_contact_level_sorting_keeps_match_tier_first(self) -> None:
        def add_candidate(external_id, score, verified_email=False):
            candidate = self.candidate(external_id)
            candidate["external_id"] = external_id
            candidate["username"] = external_id
            candidate["display_name"] = external_id
            candidate["match_score"] = score
            candidate["profile_url"] = "https://profiles.example.test/{}".format(
                external_id
            )
            candidate["contact_url"] = candidate["profile_url"]
            candidate["contact_email"] = ""
            if verified_email:
                candidate["contact_email"] = "{}@example.test".format(external_id)
                candidate["contact_email_source_url"] = candidate["profile_url"]
                candidate["contact_email_verified_at"] = (
                    "2026-08-04T10:00:00+08:00"
                )
            db.upsert_candidate(candidate)

        add_candidate("high-profile-only", 82)
        add_candidate("medium-profile-only", 81)
        add_candidate("medium-email", 70, verified_email=True)
        add_candidate("low-email", 69, verified_email=True)

        ordered = [
            candidate["external_id"] for candidate in db.list_candidates({})["items"]
        ]
        self.assertEqual(
            ordered,
            [
                "high-profile-only",
                "medium-email",
                "medium-profile-only",
                "low-email",
            ],
        )

    def test_contactability_filters_cover_email_other_entry_and_profile(self) -> None:
        def add_candidate(external_id, level):
            candidate = self.candidate(external_id)
            candidate["external_id"] = external_id
            candidate["username"] = external_id
            candidate["profile_url"] = "https://profiles.example.test/{}".format(
                external_id
            )
            candidate["contact_url"] = candidate["profile_url"]
            candidate["contact_email"] = ""
            if level in {"A", "B"}:
                candidate["contact_email"] = "{}@example.test".format(external_id)
            if level == "A":
                candidate["contact_email_source_url"] = candidate["profile_url"]
                candidate["contact_email_verified_at"] = (
                    "2026-08-04T10:00:00+08:00"
                )
            if level == "C":
                candidate["contact_url"] = "https://contact.example.test/{}".format(
                    external_id
                )
            candidate_id, _ = db.upsert_candidate(candidate)
            self.assertEqual(db.get_candidate(candidate_id)["contact_level"], level)

        for level in ("A", "B", "C", "D"):
            add_candidate("level-{}".format(level.lower()), level)

        self.assertEqual(db.list_candidates({"contactability": "email"})["total"], 2)
        self.assertEqual(
            db.list_candidates({"contactability": "contactable"})["total"], 3
        )
        self.assertEqual(
            db.list_candidates({"contactability": "profile_only"})["total"], 1
        )
        for level in ("A", "B", "C", "D"):
            result = db.list_candidates({"contactability": level})
            self.assertEqual(result["total"], 1)
            self.assertEqual(result["items"][0]["contact_level"], level)
        with self.assertRaises(ValueError):
            db.list_candidates({"contactability": "unknown"})

    def test_public_email_validation_rejects_noreply(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        with self.assertRaises(ValueError):
            db.set_public_email(
                candidate_id,
                "123@users.noreply.github.com",
                "https://github.com/candidate",
            )

    def test_export_includes_candidate_evidence(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        exported = db.export_candidates()
        self.assertEqual(exported[0]["id"], candidate_id)
        self.assertEqual(exported[0]["evidence"][0]["title"], "agent-kit")

    def test_archive_hides_candidate_and_restore_returns_it(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        self.assertTrue(db.archive_candidate(candidate_id))
        self.assertEqual(db.list_candidates({})["total"], 0)
        archived = db.list_candidates({"archived": "only"})
        self.assertEqual(archived["total"], 1)
        self.assertTrue(archived["items"][0]["archived_at"])
        self.assertEqual(db.export_candidates(), [])
        self.assertTrue(db.restore_candidate(candidate_id))
        self.assertEqual(db.list_candidates({})["total"], 1)

    def test_permanent_delete_only_accepts_archived_candidate(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        self.assertFalse(db.delete_archived_candidate(candidate_id))
        db.archive_candidate(candidate_id)
        self.assertTrue(db.delete_archived_candidate(candidate_id))
        self.assertIsNone(db.get_candidate(candidate_id))

    def test_backup_and_management_stats_use_local_database(self) -> None:
        db.upsert_candidate(self.candidate())
        backup = db.create_backup()
        backup_path = Path(self.temp_dir.name) / "backups" / backup["filename"]
        self.assertTrue(backup_path.is_file())
        self.assertGreater(backup["size_bytes"], 0)
        stats = db.data_management_stats()
        self.assertEqual(stats["active_candidates"], 1)
        self.assertEqual(stats["archived_candidates"], 0)
        self.assertEqual(stats["backup_count"], 1)
        self.assertEqual(stats["latest_backup"]["filename"], backup["filename"])

    def test_archive_nonmatching_and_cleanup_old_jobs(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        db.review_candidate(candidate_id, "不符合", "不匹配")
        self.assertEqual(db.archive_nonmatching_candidates(), 1)
        job_id = db.create_job("测试", {"mode": "search"})
        with db.connect() as connection:
            connection.execute(
                "UPDATE jobs SET status = '已完成', created_at = ? WHERE id = ?",
                ("2025-01-01T00:00:00+08:00", job_id),
            )
        self.assertEqual(db.cleanup_job_logs(90), 1)

    def test_rejects_invalid_verification_value(self) -> None:
        candidate_id, _ = db.upsert_candidate(self.candidate())
        with self.assertRaises(ValueError):
            db.review_candidate(
                candidate_id,
                "优先联系",
                "",
                education_verification="博士",
            )

    def test_upsert_rejects_unsafe_profile_and_skips_unsafe_evidence(self) -> None:
        unsafe = self.candidate()
        unsafe["profile_url"] = "javascript:alert(1)"
        with self.assertRaises(ValueError):
            db.upsert_candidate(unsafe)

        safe_candidate = self.candidate("Updated")
        safe_candidate["external_id"] = "456"
        candidate_id, _ = db.upsert_candidate(
            {
                **safe_candidate,
                "evidence": [
                    {"title": "unsafe", "url": "data:text/html,invalid"},
                    {"title": "safe", "url": "https://example.com/safe"},
                ],
            }
        )
        saved = db.get_candidate(candidate_id)
        self.assertEqual([item["title"] for item in saved["evidence"]], ["safe"])

    def test_restart_marks_pending_and_cancel_requested_jobs_interrupted(self) -> None:
        waiting = db.create_job("等待", {})
        cancelling = db.create_job("取消", {})
        with db.connect() as connection:
            connection.execute("UPDATE jobs SET status = '请求取消' WHERE id = ?", (cancelling,))
        db.init_db()
        jobs = {item["id"]: item for item in db.list_jobs(10)}
        self.assertEqual(jobs[waiting]["status"], "执行中断")
        self.assertEqual(jobs[cancelling]["status"], "执行中断")


class JobSelectionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.previous_db = os.environ.get("TALENT_RADAR_DB")
        os.environ["TALENT_RADAR_DB"] = str(
            Path(self.temp_dir.name) / "job-selection.db"
        )
        db.init_db()

    def tearDown(self) -> None:
        if self.previous_db is None:
            os.environ.pop("TALENT_RADAR_DB", None)
        else:
            os.environ["TALENT_RADAR_DB"] = self.previous_db
        self.temp_dir.cleanup()

    @staticmethod
    def candidate(source, external_id, score, verified_email=False):
        profile_url = "https://profiles.example.test/{}/{}".format(
            source, external_id
        )
        candidate = {
            "source": source,
            "external_id": external_id,
            "username": external_id,
            "display_name": external_id,
            "city": "北京",
            "profile_url": profile_url,
            "contact_url": profile_url,
            "contact_email": "",
            "suggested_role": "AI Agent 工程师",
            "match_score": score,
            "evidence": [],
        }
        if verified_email:
            candidate["contact_email"] = "{}@example.test".format(external_id)
            candidate["contact_email_source_url"] = profile_url
            candidate["contact_email_verified_at"] = (
                "2026-08-04T10:00:00+08:00"
            )
        return candidate

    def run_selection(self, prefer_contactable):
        github_collector = Mock(
            return_value=[
                self.candidate("github", "github-top", 80),
                self.candidate("github", "github-second", 79),
            ]
        )
        gitee_collector = Mock(
            return_value=[
                self.candidate("gitee", "gitee-contact", 75, verified_email=True)
            ]
        )
        config = normalize_config(
            {
                "target": 1,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github", "gitee"],
                "prefer_contactable": prefer_contactable,
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch.dict(
            jobs.SEARCH_COLLECTORS,
            {"github": github_collector, "gitee": gitee_collector},
        ), patch.dict(os.environ, {"GITHUB_TOKEN": ""}):
            JobManager()._run_search_job(job_id, config)
        return db.get_job(job_id), github_collector, gitee_collector

    def test_contact_priority_selects_across_all_configured_sources(self) -> None:
        job, github_collector, gitee_collector = self.run_selection(True)

        self.assertEqual(job["status"], "已完成")
        self.assertEqual(job["result_count"], 1)
        self.assertEqual(job["candidates"][0]["username"], "gitee-contact")
        expected_keywords = ["agent", "langgraph", "智能体"]
        self.assertEqual(
            [call.args[0] for call in github_collector.call_args_list],
            expected_keywords,
        )
        self.assertEqual(
            [call.args[0] for call in gitee_collector.call_args_list],
            expected_keywords,
        )

    def test_contact_priority_can_be_disabled(self) -> None:
        job, _, gitee_collector = self.run_selection(False)

        self.assertEqual(job["candidates"][0]["username"], "github-top")
        self.assertEqual(
            [call.args[0] for call in gitee_collector.call_args_list],
            ["agent", "langgraph", "智能体"],
        )

    def test_existing_verified_email_is_reused_before_selection(self) -> None:
        known = self.candidate("gitee", "known-contact", 75)
        candidate_id, _ = db.upsert_candidate(known)
        db.set_public_email(
            candidate_id,
            "known-contact@example.test",
            known["profile_url"],
        )
        github_collector = Mock(
            return_value=[self.candidate("github", "higher-score", 80)]
        )
        gitee_collector = Mock(return_value=[known])
        config = normalize_config(
            {
                "target": 1,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github", "gitee"],
                "prefer_contactable": True,
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch.dict(
            jobs.SEARCH_COLLECTORS,
            {"github": github_collector, "gitee": gitee_collector},
        ), patch.dict(os.environ, {"GITHUB_TOKEN": ""}):
            JobManager()._run_search_job(job_id, config)

        job = db.get_job(job_id)
        self.assertEqual(job["candidates"][0]["username"], "known-contact")
        self.assertEqual(job["candidates"][0]["contact_level"], "A")
        self.assertIn("目标 1", job["message"])
        self.assertIn("实际 1", job["message"])
        self.assertIn("新增 0", job["message"])
        self.assertIn("已有 1", job["message"])
        self.assertIn("公开联系方式 1", job["message"])

    def test_job_log_reports_target_actual_dedup_and_insert_counts(self) -> None:
        candidates = [self.candidate("github", "candidate-{}".format(index), 70 - index) for index in range(8)]
        collector = Mock(side_effect=[candidates, [], []])
        config = normalize_config(
            {
                "target": 10,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
                "prefer_contactable": True,
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch.dict(jobs.SEARCH_COLLECTORS, {"github": collector}), patch.dict(
            os.environ, {"GITHUB_TOKEN": "token-for-test"}
        ):
            JobManager()._run_search_job(job_id, config)
        job = db.get_job(job_id)
        self.assertEqual(job["target_count"], 10)
        self.assertEqual(job["result_count"], 8)
        self.assertEqual(job["discovered_count"], 8)
        self.assertEqual(job["unique_count"], 8)
        self.assertEqual(job["duplicate_count"], 0)
        self.assertEqual(job["filtered_count"], 0)
        self.assertEqual(job["inserted_count"], 8)
        self.assertEqual(job["existing_count"], 0)
        self.assertIn("目标 10", job["message"])
        self.assertIn("实际 8", job["message"])
        self.assertIn("缺口 2", job["message"])

    def test_job_log_distinguishes_existing_candidates_and_task_duplicates(self) -> None:
        existing = self.candidate("github", "existing", 80)
        db.upsert_candidate(existing)
        new = self.candidate("github", "new", 79)
        collector = Mock(side_effect=[[existing, new, existing], [], []])
        config = normalize_config(
            {
                "target": 2,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch.dict(jobs.SEARCH_COLLECTORS, {"github": collector}), patch.dict(
            os.environ, {"GITHUB_TOKEN": "token-for-test"}
        ):
            JobManager()._run_search_job(job_id, config)
        job = db.get_job(job_id)
        self.assertEqual(job["result_count"], 2)
        self.assertEqual(job["discovered_count"], 3)
        self.assertEqual(job["unique_count"], 2)
        self.assertEqual(job["duplicate_count"], 1)
        self.assertEqual(job["inserted_count"], 1)
        self.assertEqual(job["existing_count"], 1)
        self.assertIn("任务内去重 1", job["message"])
        self.assertIn("新增 1", job["message"])
        self.assertIn("已有 1", job["message"])

    def test_job_log_counts_failed_source_calls_without_losing_successful_results(self) -> None:
        successful = self.candidate("github", "reachable", 80)
        github = Mock(side_effect=[[successful], [], []])
        gitee = Mock(side_effect=collectors.NetworkUnavailable("VPN unavailable"))
        config = normalize_config(
            {
                "target": 1,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github", "gitee"],
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch.dict(jobs.SEARCH_COLLECTORS, {"github": github, "gitee": gitee}), patch.dict(
            os.environ, {"GITHUB_TOKEN": "test-token"}
        ):
            JobManager()._run_search_job(job_id, config)
        job = db.get_job(job_id)
        self.assertEqual(job["result_count"], 1)
        self.assertEqual(job["source_success_count"], 3)
        self.assertEqual(job["source_failure_count"], 3)
        self.assertEqual(job["source_stats"][1]["failures"], 3)
        self.assertIn("来源失败 3", job["message"])
        self.assertIn("Gitee", job["error"])

    def test_url_failure_persists_one_target_and_source_reason(self) -> None:
        config = normalize_config(
            {
                "mode": "url",
                # The submitted search target may still be 10, but URL mode
                # represents exactly one requested public profile.
                "target": 10,
                "url": "https://github.com/example-user",
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
                "use_local_ai": False,
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch(
            "jobs.analyze_public_url",
            side_effect=collectors.NetworkUnavailable("offline"),
        ):
            JobManager()._run_job(job_id, "手动采集", config)

        job = db.get_job(job_id)
        self.assertEqual(job["status"], "网络不可用")
        self.assertEqual(job["target_count"], 1)
        self.assertEqual(job["result_count"], 0)
        self.assertEqual(job["source_failure_count"], 1)
        self.assertEqual(job["source_stats"][0]["source"], "github")
        self.assertEqual(job["source_stats"][0]["attempts"], 1)
        self.assertEqual(job["source_stats"][0]["failures"], 1)
        self.assertEqual(job["source_stats"][0]["errors"], ["offline"])
        self.assertIn("目标 1", job["message"])
        self.assertIn("实际 0", job["message"])
        self.assertIn("来源失败 1", job["message"])

    def test_save_failure_keeps_completed_batch_metrics(self) -> None:
        candidates = [
            self.candidate("github", "partial-1", 80),
            self.candidate("github", "partial-2", 79),
        ]
        collector = Mock(side_effect=[candidates, [], []])
        config = normalize_config(
            {
                "target": 2,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
                "use_local_ai": False,
            }
        )
        job_id = db.create_job("手动采集", config)
        original_upsert = db.upsert_candidate
        calls = {"count": 0}

        def fail_on_second(candidate, job_id=None):
            calls["count"] += 1
            if calls["count"] == 2:
                raise RuntimeError("synthetic save failure")
            return original_upsert(candidate, job_id=job_id)

        with patch.dict(jobs.SEARCH_COLLECTORS, {"github": collector}), patch.dict(
            os.environ, {"GITHUB_TOKEN": "test-token"}
        ), patch("jobs.db.upsert_candidate", side_effect=fail_on_second):
            with self.assertRaisesRegex(RuntimeError, "synthetic save failure"):
                JobManager()._run_job(job_id, "手动采集", config)

        job = db.get_job(job_id)
        self.assertEqual(job["status"], "执行失败")
        self.assertEqual(job["target_count"], 2)
        self.assertEqual(job["discovered_count"], 2)
        self.assertEqual(job["unique_count"], 2)
        self.assertEqual(job["result_count"], 1)
        self.assertEqual(job["inserted_count"], 1)
        self.assertEqual(job["existing_count"], 0)
        self.assertIn("目标 2", job["message"])
        self.assertIn("实际 1", job["message"])
        self.assertIn("新增 1", job["message"])
        self.assertIn("保存阶段异常", job["message"])

    def test_url_cancel_does_not_report_unsaved_profile_as_filtered(self) -> None:
        candidate = self.candidate("github", "cancelled-url", 80)
        config = normalize_config(
            {
                "mode": "url",
                "target": 10,
                "url": "https://github.com/cancelled-url",
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
                "use_local_ai": False,
            }
        )
        job_id = db.create_job("手动采集", config)
        with patch("jobs.analyze_public_url", return_value=candidate), patch(
            "jobs.db.job_cancel_requested", return_value=True
        ):
            JobManager()._run_job(job_id, "手动采集", config)

        job = db.get_job(job_id)
        self.assertEqual(job["status"], "已取消")
        self.assertEqual(job["target_count"], 1)
        self.assertEqual(job["result_count"], 0)
        self.assertEqual(job["filtered_count"], 0)
        self.assertIn("取消未保存 1", job["message"])

    def test_job_log_reports_completed_local_ai_analysis(self) -> None:
        collector = Mock(side_effect=[[self.candidate("github", "ai-ready", 70)], [], []])
        config = normalize_config(
            {
                "target": 1,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
                "use_local_ai": True,
            }
        )
        ai_result = {
            "match_score": 90,
            "confidence": 0.8,
            "summary": "strong public project evidence",
            "evidence": [],
            "model": "qwen3:4b",
            "matched_at": "2026-08-17T10:00:00+08:00",
        }
        job_id = db.create_job("手动采集", config)
        with patch.dict(jobs.SEARCH_COLLECTORS, {"github": collector}), patch.dict(
            os.environ, {"GITHUB_TOKEN": "test-token"}
        ), patch("jobs.match_candidate", return_value=ai_result):
            JobManager()._run_search_job(job_id, config)
        job = db.get_job(job_id)
        self.assertTrue(job["ai_requested"])
        self.assertEqual(job["ai_completed_count"], 1)
        self.assertEqual(job["ai_fallback_count"], 0)
        self.assertEqual(job["candidates"][0]["ai_match_status"], "已完成")
        self.assertIn("AI 完成 1", job["message"])
        self.assertIn("AI 回退 0", job["message"])

    def test_one_ai_runtime_failure_falls_back_once_for_the_whole_job(self) -> None:
        candidates = [
            self.candidate("github", "ai-offline-1", 70),
            self.candidate("github", "ai-offline-2", 69),
        ]
        collector = Mock(side_effect=[candidates, [], []])
        config = normalize_config(
            {
                "target": 2,
                "roles": ["AI Agent 工程师"],
                "cities": ["北京"],
                "sources": ["github"],
                "use_local_ai": True,
            }
        )
        job_id = db.create_job("手动采集", config)
        matcher = Mock(side_effect=jobs.OllamaUnavailable("local model offline"))
        with patch.dict(jobs.SEARCH_COLLECTORS, {"github": collector}), patch.dict(
            os.environ, {"GITHUB_TOKEN": "test-token"}
        ), patch("jobs.match_candidate", matcher):
            JobManager()._run_search_job(job_id, config)
        job = db.get_job(job_id)
        self.assertEqual(matcher.call_count, 1)
        self.assertEqual(job["ai_completed_count"], 0)
        self.assertEqual(job["ai_fallback_count"], 2)
        self.assertTrue(all(item["ai_match_status"] == "不可用，已回退规则" for item in job["candidates"]))
        self.assertIn("AI 回退 2", job["message"])
        self.assertIn("AI 原因 local model offline", job["message"])


class AiMetricsTests(unittest.TestCase):
    def test_disabled_ai_does_not_count_historical_completed_status(self) -> None:
        metrics = jobs._ai_counts(
            [{"ai_match_status": "已完成"}, {"ai_match_status": "不可用，已回退规则"}],
            requested=False,
        )
        self.assertEqual(metrics["ai_requested"], 0)
        self.assertEqual(metrics["ai_completed_count"], 0)
        self.assertEqual(metrics["ai_fallback_count"], 0)
        self.assertEqual(metrics["ai_disabled_count"], 2)


class ExcelLinkTests(unittest.TestCase):
    def test_excel_text_escapes_formula_prefixes(self) -> None:
        self.assertEqual(excel_export._excel_safe_value("=HYPERLINK(\"x\")"), "'=HYPERLINK(\"x\")")
        self.assertEqual(excel_export._excel_safe_value("candidate"), "candidate")

    def test_excel_text_removes_xml_illegal_control_characters(self) -> None:
        self.assertEqual(excel_export._excel_text("agent\x00profile\x07"), "agentprofile")
        self.assertEqual(excel_export._excel_text("line one\nline two"), "line one\nline two")

    def test_export_link_validation_handles_malformed_urls_and_preserves_email_case(self) -> None:
        self.assertEqual(excel_export._safe_external_url("http://[broken"), "")
        self.assertEqual(excel_export._safe_external_url("https://example.test/a b"), "")
        self.assertEqual(
            excel_export._safe_mailto("Candidate@Example.TEST"),
            "mailto:Candidate@Example.TEST",
        )

    def test_excel_datetime_converts_utc_to_shanghai(self) -> None:
        self.assertEqual(
            excel_export._excel_shanghai_datetime("2026-08-04T02:03:00Z"),
            datetime(2026, 8, 4, 10, 3),
        )
        self.assertEqual(
            excel_export._excel_shanghai_datetime("not-a-date"),
            "not-a-date",
        )

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
    def test_portable_export_preserves_sheets_fields_links_and_dates(self) -> None:
        from openpyxl import load_workbook

        candidates = [
            {
                "id": 7,
                "display_name": "Synthetic Candidate",
                "username": "synthetic-candidate",
                "source": "github",
                "city": "北京",
                "suggested_role": "AI Agent 工程师",
                "match_score": 86,
                "company": "Example University",
                "education_status": "公开资料提及本科",
                "education_verification": "本科及以上",
                "age_status": "30岁以下",
                "work_location_status": "接受北京",
                "agent_experience_status": "原创 Agent 项目",
                "review_status": "优先联系",
                "contact_stage": "已联系",
                "contact_updated_at": "2026-08-04T02:03:00Z",
                "contact_level": "A",
                "contact_email": "candidate@example.test",
                "contact_email_source_url": "https://profiles.example.test/candidate",
                "contact_email_verified_at": "2026-08-04T10:04:00+08:00",
                "profile_url": "https://profiles.example.test/candidate",
                "contact_url": "https://contact.example.test/candidate",
                "bio": "Synthetic AI agent profile",
                "review_note": "Synthetic test record",
                "first_seen_at": "2026-08-01T09:00:00+08:00",
                "last_seen_at": "2026-08-05T09:00:00+08:00",
                "updated_at": "2026-08-05T09:30:00+08:00",
                "evidence": [
                    {
                        "title": "synthetic-agent-project",
                        "url": "https://code.example.test/synthetic-agent-project",
                        "language": "Python",
                        "stars": 12,
                        "is_fork": False,
                        "description": "Synthetic public project evidence",
                    }
                ],
            }
        ]

        content = excel_export.generate_excel(candidates)
        self.assertTrue(content.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        self.addCleanup(workbook.close)
        self.assertEqual(workbook.sheetnames, ["候选人总表", "项目证据"])

        candidate_sheet = workbook["候选人总表"]
        evidence_sheet = workbook["项目证据"]
        self.assertEqual(
            [candidate_sheet.cell(6, column).value for column in range(1, 28)],
            excel_export.CANDIDATE_HEADERS,
        )
        self.assertEqual(candidate_sheet["Q7"].value, "A")
        self.assertEqual(candidate_sheet["R7"].hyperlink.target, "mailto:candidate@example.test")
        self.assertEqual(
            candidate_sheet["S7"].hyperlink.target,
            "https://profiles.example.test/candidate",
        )
        self.assertEqual(
            candidate_sheet["U7"].hyperlink.target,
            "https://profiles.example.test/candidate",
        )
        self.assertEqual(
            candidate_sheet["V7"].hyperlink.target,
            "https://contact.example.test/candidate",
        )
        self.assertEqual(candidate_sheet["P7"].value, datetime(2026, 8, 4, 10, 3))
        self.assertEqual(candidate_sheet["P7"].number_format, "yyyy-mm-dd hh:mm")
        self.assertEqual(candidate_sheet["T7"].value, datetime(2026, 8, 4, 10, 4))
        self.assertEqual(candidate_sheet["Y7"].value, datetime(2026, 8, 1, 9, 0))
        self.assertEqual(candidate_sheet["B3"].value, "=COUNTA(B7:B7)")
        self.assertEqual(candidate_sheet.freeze_panes, "C7")
        self.assertIn("CandidatesTable", candidate_sheet.tables)
        self.assertEqual(candidate_sheet["U7"].border.bottom.style, "thin")
        self.assertEqual(candidate_sheet["U7"].alignment.vertical, "top")
        self.assertTrue(
            any(
                "Q7" in validation
                for validation in candidate_sheet.data_validations.dataValidation
            )
        )

        self.assertEqual(
            [evidence_sheet.cell(3, column).value for column in range(1, 9)],
            excel_export.EVIDENCE_HEADERS,
        )
        self.assertEqual(
            evidence_sheet["D4"].hyperlink.target,
            "https://code.example.test/synthetic-agent-project",
        )
        self.assertEqual(evidence_sheet.freeze_panes, "A4")
        self.assertIn("EvidenceTable", evidence_sheet.tables)

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
    def test_portable_export_rejects_unsafe_links_and_formula_text(self) -> None:
        from openpyxl import load_workbook

        content = excel_export.generate_excel(
            [
                {
                    "id": 1,
                    "display_name": "=2+2",
                    "profile_url": "javascript:alert(1)",
                    "contact_url": "javascript:alert(2)",
                    "contact_email": "123@users.noreply.github.com",
                    "evidence": [],
                }
            ]
        )
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        self.addCleanup(workbook.close)
        sheet = workbook["候选人总表"]
        self.assertEqual(sheet["B7"].value, "'=2+2")
        self.assertIsNone(sheet["R7"].hyperlink)
        self.assertIsNone(sheet["U7"].hyperlink)
        self.assertIsNone(sheet["V7"].hyperlink)


class AppHandlerTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.previous_db = os.environ.get("TALENT_RADAR_DB")
        os.environ["TALENT_RADAR_DB"] = str(Path(self.temp_dir.name) / "test.db")
        db.init_db()
        self.server = ThreadingHTTPServer(("127.0.0.1", 0), app.AppHandler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()

    def tearDown(self) -> None:
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=2)
        if self.previous_db is None:
            os.environ.pop("TALENT_RADAR_DB", None)
        else:
            os.environ["TALENT_RADAR_DB"] = self.previous_db
        self.temp_dir.cleanup()

    def request_raw(self, method: str, path: str, payload=None, extra_headers=None):
        connection = http.client.HTTPConnection("127.0.0.1", self.server.server_port, timeout=3)
        headers = dict(extra_headers or {})
        body = None
        if payload is not None:
            headers["Content-Type"] = "application/json"
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        try:
            connection.request(method, path, body=body, headers=headers)
            response = connection.getresponse()
            raw_bytes = response.read()
            headers = dict(response.getheaders())
            return response.status, headers, raw_bytes
        finally:
            connection.close()

    def request(self, method: str, path: str, payload=None, extra_headers=None):
        status, headers, raw_bytes = self.request_raw(method, path, payload, extra_headers)
        content_type = headers.get("Content-Type", "").lower()
        is_textual = content_type.startswith("text/") or content_type.startswith(
            ("application/json", "application/javascript", "application/xml")
        )
        is_binary = bool(content_type) and not is_textual
        if is_binary:
            return status, headers, raw_bytes
        result = raw_bytes.decode("utf-8")
        parsed = json.loads(result) if result and content_type.startswith("application/json") else result
        return status, headers, parsed

    def add_candidate_at(self, external_id, collected_at, evidence_title=None):
        candidate = {
            "source": "github",
            "external_id": external_id,
            "username": external_id,
            "display_name": external_id,
            "city": "北京",
            "profile_url": "https://profiles.example.test/{}".format(external_id),
            "contact_url": "https://profiles.example.test/{}".format(external_id),
            "contact_email": "{}@example.test".format(external_id),
            "suggested_role": "AI Agent 工程师",
            "match_score": 80,
            "evidence": [
                {
                    "title": evidence_title or "{}-project".format(external_id),
                    "url": "https://code.example.test/{}/project".format(external_id),
                    "description": "synthetic agent project",
                    "stars": 1,
                    "is_fork": False,
                }
            ],
        }
        with patch.object(db, "now_iso", return_value=collected_at):
            candidate_id, _ = db.upsert_candidate(candidate)
        return candidate_id

    def test_invalid_query_returns_json_error_and_security_header(self) -> None:
        status, _, payload = self.request("GET", "/api/jobs?limit=bad")
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        status, headers, _ = self.request("GET", "/")
        self.assertEqual(status, 200)
        self.assertEqual(headers["X-Frame-Options"], "DENY")

    def test_invalid_candidate_date_returns_400_on_list(self) -> None:
        status, _, payload = self.request(
            "GET", "/api/candidates?last_seen_from=2026-02-30"
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertEqual(payload["error"], "开始日期格式无效，请使用 YYYY-MM-DD")

    def test_candidate_date_query_filters_list_endpoint(self) -> None:
        self.add_candidate_at("in-range", "2026-08-18T10:00:00+08:00")
        self.add_candidate_at("outside", "2026-08-17T10:00:00+08:00")

        status, _, payload = self.request(
            "GET", "/api/candidates?last_seen_from=2026-08-18&last_seen_to=2026-08-18"
        )

        self.assertEqual(status, 200)
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["items"][0]["external_id"], "in-range")

    def test_invalid_candidate_date_returns_400_on_export(self) -> None:
        status, headers, payload = self.request(
            "GET", "/export/candidates.xlsx?last_seen_to=2101-01-01"
        )

        self.assertEqual(status, 400)
        self.assertTrue(headers["Content-Type"].startswith("application/json"))
        self.assertIn("error", payload)

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
    def test_candidate_excel_export_applies_date_query(self) -> None:
        from openpyxl import load_workbook

        self.add_candidate_at(
            "in-range",
            "2026-08-18T09:00:00+08:00",
            evidence_title="in-range-evidence",
        )
        self.add_candidate_at(
            "outside",
            "2026-08-17T09:00:00+08:00",
            evidence_title="outside-evidence",
        )

        status, headers, content = self.request(
            "GET",
            "/export/candidates.xlsx?last_seen_from=2026-08-18&last_seen_to=2026-08-18",
        )

        self.assertEqual(status, 200)
        self.assertTrue(
            headers["Content-Type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        self.addCleanup(workbook.close)
        candidate_names = [
            cell.value
            for cell in workbook["候选人总表"]["B"][6:]
            if cell.value is not None
        ]
        evidence_titles = [
            cell.value
            for cell in workbook["项目证据"]["C"][3:]
            if cell.value is not None
        ]
        self.assertEqual(candidate_names, ["in-range"])
        self.assertEqual(evidence_titles, ["in-range-evidence"])

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is not installed")
    def test_candidate_excel_export_supports_empty_date_result(self) -> None:
        from openpyxl import load_workbook

        self.add_candidate_at("outside", "2026-08-17T09:00:00+08:00")

        status, _, content = self.request(
            "GET",
            "/export/candidates.xlsx?last_seen_from=2026-08-18&last_seen_to=2026-08-18",
        )

        self.assertEqual(status, 200)
        workbook = load_workbook(io.BytesIO(content), data_only=False)
        self.addCleanup(workbook.close)
        candidate_sheet = workbook["候选人总表"]
        self.assertEqual(candidate_sheet["B6"].value, "候选人")
        self.assertIsNone(candidate_sheet["B7"].value)

    def test_source_health_endpoint_returns_static_catalog_without_probe(self) -> None:
        status, _, payload = self.request("GET", "/api/source-health")
        self.assertEqual(status, 200)
        self.assertEqual(len(payload["items"]), 15)
        statuses = {item["key"]: item["status"] for item in payload["items"]}
        self.assertEqual(statuses["github"], "not_checked")
        self.assertEqual(statuses["modelscope"], "not_implemented")

    def test_missing_candidate_patch_returns_not_found(self) -> None:
        status, _, payload = self.request(
            "PATCH",
            "/api/candidates/999",
            {"review_status": "待审核"},
        )
        self.assertEqual(status, 404)
        self.assertEqual(payload["error"], "候选人不存在")

    def test_write_accepts_loopback_origin_from_forwarded_host_port(self) -> None:
        forwarded_port = self.server.server_port + 1
        status, _, payload = self.request(
            "PATCH",
            "/api/candidates/999",
            {"review_status": "待审核"},
            {
                "Host": "127.0.0.1:{}".format(forwarded_port),
                "Origin": "http://127.0.0.1:{}".format(forwarded_port),
            },
        )

        self.assertEqual(status, 404)
        self.assertEqual(payload["error"], "候选人不存在")

    def test_write_rejects_non_loopback_forwarded_host_origin(self) -> None:
        status, _, payload = self.request(
            "PATCH",
            "/api/candidates/999",
            {"review_status": "待审核"},
            {"Host": "example.test:18767", "Origin": "http://example.test:18767"},
        )

        self.assertEqual(status, 400)
        self.assertEqual(payload["error"], "请求来源无效")

    def test_schedule_request_is_normalized_at_api_boundary(self) -> None:
        status, _, payload = self.request(
            "PUT",
            "/api/schedule",
            {
                "enabled": False,
                "weekday": 0,
                "hour": 9,
                "minute": 0,
                "config": {
                    "target": 999,
                    "roles": ["无效岗位"],
                    "cities": ["北京"],
                    "sources": ["github"],
                    "prefer_contactable": False,
                },
            },
        )
        self.assertEqual(status, 200)
        self.assertEqual(payload["config"]["target"], 50)
        self.assertEqual(payload["config"]["roles"], ["AI Agent 工程师"])
        self.assertFalse(payload["config"]["prefer_contactable"])

    def test_schedule_rejects_non_boolean_contact_priority(self) -> None:
        status, _, payload = self.request(
            "PUT",
            "/api/schedule",
            {
                "enabled": False,
                "weekday": 0,
                "hour": 10,
                "minute": 0,
                "config": {"prefer_contactable": "false"},
            },
        )
        self.assertEqual(status, 400)
        self.assertIn("布尔值", payload["error"])


if __name__ == "__main__":
    unittest.main()
